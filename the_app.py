import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

# Load the data from the Excel files
agent_list = pd.read_excel(r"C:\Users\marcus.forsen\Desktop\new project\agents.xlsx")
call_results_conversion = pd.read_excel(r"C:\Users\marcus.forsen\Desktop\new project\Agent_Call_Results.xlsx", sheet_name='Conversion Agents')
call_results_retention = pd.read_excel(r"C:\Users\marcus.forsen\Desktop\new project\Agent_Call_Results.xlsx", sheet_name='Retention Agents')
duration_results_conversion = pd.read_excel(r"C:\Users\marcus.forsen\Desktop\new project\Agent_Duration_Results.xlsx", sheet_name='Conversion Agents')
duration_results_retention = pd.read_excel(r"C:\Users\marcus.forsen\Desktop\new project\Agent_Duration_Results.xlsx", sheet_name='Retention Agents')

# Clean and standardize the 'Agent Name' in all dataframes
agent_list['AGENTNAME'] = agent_list['AGENTNAME'].str.strip().str.upper()
call_results_conversion['Agent Name'] = call_results_conversion['Agent Name'].str.strip().str.upper()
call_results_retention['Agent Name'] = call_results_retention['Agent Name'].str.strip().str.upper()
duration_results_conversion['Agent Name'] = duration_results_conversion['Agent Name'].str.strip().str.upper()
duration_results_retention['Agent Name'] = duration_results_retention['Agent Name'].str.strip().str.upper()

# Extract and rename columns from the agent list
agents_data = agent_list[['AGENTNAME', 'DESK', 'DEPARTMENT']].copy()
agents_data.rename(columns={'AGENTNAME': 'Agent Name', 'DESK': 'Desk'}, inplace=True)

# Define the custom desk order for "Conversion Agents"
desk_order_conversion = [
    'Team Elly', 'Team Vincent', 'Team Rahul', 'Team Sameer', 'Team Eden', 'Team Elena', 'Team Larisa'
]
desk_order_conversion_dict = {desk: idx for idx, desk in enumerate(desk_order_conversion)}

# Define the custom desk order for "Retention Agents"
desk_order_retention = [
    'Japan Team', 'Korean Team', 'Aarav Team', 'Ajay Team', 'French Maxime', 
    'AKA Team', 'Spanish Andres', 'Portuguese Pedro'
]
desk_order_retention_dict = {desk: idx for idx, desk in enumerate(desk_order_retention)}

# Extract relevant columns and handle call results and durations for Conversion Agents
call_data_conversion = call_results_conversion[['Agent Name', 'Unique', 'Call Attempts']].copy()
duration_data_conversion = duration_results_conversion[['Agent Name', 'Total Time']].copy()

# Convert the "Total Time" to timedelta and handle formatting
duration_data_conversion['Total Time'] = pd.to_timedelta(duration_data_conversion['Total Time'], errors='coerce')
duration_data_conversion['Duration'] = duration_data_conversion['Total Time'].apply(lambda x: str(x).split()[-1] if pd.notnull(x) else '0')
duration_data_conversion.drop(columns=['Total Time'], inplace=True)

# Extract relevant columns and handle call results and durations for Retention Agents
call_data_retention = call_results_retention[['Agent Name', 'Unique', 'Call Attempts']].copy()
duration_data_retention = duration_results_retention[['Agent Name', 'Total Time']].copy()

# Convert the "Total Time" to timedelta and handle formatting
duration_data_retention['Total Time'] = pd.to_timedelta(duration_data_retention['Total Time'], errors='coerce')
duration_data_retention['Duration'] = duration_data_retention['Total Time'].apply(lambda x: str(x).split()[-1] if pd.notnull(x) else '0')
duration_data_retention.drop(columns=['Total Time'], inplace=True)

# Merge the "Unique", "Call Attempts" and "Duration" columns with the agent data for Conversion Agents
merged_data_conversion = pd.merge(agents_data[agents_data['DEPARTMENT'] == 1].copy(), call_data_conversion, on='Agent Name', how='left')
merged_data_conversion = pd.merge(merged_data_conversion, duration_data_conversion, on='Agent Name', how='left')

# Merge the "Unique", "Call Attempts" and "Duration" columns with the agent data for Retention Agents
merged_data_retention = pd.merge(agents_data[agents_data['DEPARTMENT'] == 2].copy(), call_data_retention, on='Agent Name', how='left')
merged_data_retention = pd.merge(merged_data_retention, duration_data_retention, on='Agent Name', how='left')

# Fill missing values in "Unique", "Call Attempts", and "Duration" with 0
merged_data_conversion['Unique'] = merged_data_conversion['Unique'].fillna(0)
merged_data_conversion['Call Attempts'] = merged_data_conversion['Call Attempts'].fillna(0)
merged_data_conversion['Duration'] = merged_data_conversion['Duration'].fillna('0')
merged_data_conversion['Target'] = 0  # Default target

merged_data_retention['Unique'] = merged_data_retention['Unique'].fillna(0)
merged_data_retention['Call Attempts'] = merged_data_retention['Call Attempts'].fillna(0)
merged_data_retention['Duration'] = merged_data_retention['Duration'].fillna('0')
merged_data_retention['Target'] = 0  # Default target

# Define a function to calculate the Target as a percentage
def calculate_target(row):
    # Calculate percentage from unique calls
    unique_calls_percentage = (row['Unique'] // 3) * 1

    # Calculate percentage from duration
    duration_str = str(row['Duration'])
    if duration_str != '0':
        duration_parts = duration_str.split(':')
        if len(duration_parts) == 3:
            hours, minutes, seconds = map(int, duration_parts)
        elif len(duration_parts) == 2:
            hours = 0
            minutes, seconds = map(int, duration_parts)
        else:
            hours = 0
            minutes = 0
            seconds = int(duration_parts[0])
        total_minutes = hours * 60 + minutes + (seconds / 60)
        duration_percentage = (total_minutes // 1.5) * 1
    else:
        duration_percentage = 0

    # Combine the percentages
    return unique_calls_percentage + duration_percentage

# Apply the calculation function to each row
merged_data_conversion['Target'] = merged_data_conversion.apply(calculate_target, axis=1)
merged_data_retention['Target'] = merged_data_retention.apply(calculate_target, axis=1)

# Convert the Target to a numeric value for sorting
merged_data_conversion['Target Numeric'] = merged_data_conversion['Target']
merged_data_retention['Target Numeric'] = merged_data_retention['Target']

# Apply the custom sorting order for "Conversion Agents"
merged_data_conversion['Desk Order'] = merged_data_conversion['Desk'].map(desk_order_conversion_dict)
merged_data_conversion.sort_values(by=['Desk Order', 'Target Numeric'], ascending=[True, False], inplace=True)

# Apply the custom sorting order for "Retention Agents"
merged_data_retention['Desk Order'] = merged_data_retention['Desk'].map(desk_order_retention_dict)
merged_data_retention.sort_values(by=['Desk Order', 'Target Numeric'], ascending=[True, False], inplace=True)

# Drop the temporary 'Desk Order' column used for sorting
merged_data_conversion.drop(columns=['Desk Order', 'Target Numeric', 'DEPARTMENT'], inplace=True)
merged_data_retention.drop(columns=['Desk Order', 'Target Numeric', 'DEPARTMENT'], inplace=True)

# Reorder columns for Conversion Agents
merged_data_conversion = merged_data_conversion[['Desk', 'Agent Name', 'Duration', 'Call Attempts', 'Unique', 'Target']]

# Reorder columns for Retention Agents
merged_data_retention = merged_data_retention[['Desk', 'Agent Name', 'Duration', 'Call Attempts', 'Unique', 'Target']]

# Convert 'Target' column to string with percentage format
merged_data_conversion['Target'] = merged_data_conversion['Target'].astype(int).astype(str) + '%'
merged_data_retention['Target'] = merged_data_retention['Target'].astype(int).astype(str) + '%'

# Create the new Excel file with updated data
file_path = 'Agent_Results.xlsx'
with pd.ExcelWriter(file_path) as writer:
    merged_data_conversion.to_excel(writer, sheet_name='Conversion Agents', index=False)
    merged_data_retention.to_excel(writer, sheet_name='Retention Agents', index=False)

# Load the Excel file with openpyxl
wb = load_workbook(file_path)

# Get the worksheets
ws_conversion = wb['Conversion Agents']
ws_retention = wb['Retention Agents']

# Define colors for each desk
desk_colors_conversion = {
    'Team Vincent': 'FFEBEE',     # Very light red
    'Team Elena': 'E8F0FE',      # Very light blue
    'Team Eden': 'E8F0FE',      # Very light blue
    'Team Larisa': 'FFEDD5',     # Light peach
    'Team Rahul': 'FAE1C4',      # Very light beige
    'Team Sameer': 'F4BD7D',      # Very light beige
    'Team Elly': 'FFCCCC',        # Very light coral
    'Team Myles': '1452B5'        # Very light coral
}

desk_colors_retention = {
    'Aarav Team': 'FAE1C4',    # Very light beige
    'Ajay Team': 'FFF9DB',     # Light gold yellow
    'Japan Team': 'FFEBEE',    # Very light red
    'Korean Team': 'FFE4E1',   # Very light pink
    'AKA Team': 'E6FFE6',      # Very light green
    'French Maxime': 'E8F0FE', # Very light blue
    'Spanish Andres': 'FFEDD5',# Light peach
    'Portuguese Pedro': 'D4EFD4'  # Very light green
}

# Define grey color for highlighting special values
grey_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

# Apply coloring to Conversion Agents sheet
for row in ws_conversion.iter_rows(min_row=2, max_row=ws_conversion.max_row, min_col=1, max_col=6):
    desk = row[0].value
    if desk in desk_colors_conversion:
        fill = PatternFill(start_color=desk_colors_conversion[desk], end_color=desk_colors_conversion[desk], fill_type='solid')
        for cell in row:
            cell.fill = fill
            # Apply grey background for special values
            if cell.value in ['00:00:00', 0, '0%', '0']:
                cell.fill = grey_fill

# Apply coloring to Retention Agents sheet
for row in ws_retention.iter_rows(min_row=2, max_row=ws_retention.max_row, min_col=1, max_col=6):
    desk = row[0].value
    if desk in desk_colors_retention:
        fill = PatternFill(start_color=desk_colors_retention[desk], end_color=desk_colors_retention[desk], fill_type='solid')
        for cell in row:
            cell.fill = fill
            # Apply grey background for special values
            if cell.value in ['00:00:00', 0, '0%', '0']:
                cell.fill = grey_fill

# Define the thin border style
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Apply the border to Conversion Agents sheet
for row in ws_conversion.iter_rows(min_row=1, max_row=ws_conversion.max_row, min_col=1, max_col=6):
    for cell in row:
        cell.border = thin_border

# Apply the border to Retention Agents sheet
for row in ws_retention.iter_rows(min_row=1, max_row=ws_retention.max_row, min_col=1, max_col=6):
    for cell in row:
        cell.border = thin_border

# Save the updated Excel file
wb.save(file_path)

# Identify and print unmatched agents with their file source
all_agents = set(agent_list['AGENTNAME'])
call_agents_conversion = set(call_results_conversion['Agent Name'])
call_agents_retention = set(call_results_retention['Agent Name'])
duration_agents_conversion = set(duration_results_conversion['Agent Name'])
duration_agents_retention = set(duration_results_retention['Agent Name'])

unmatched_conversion = (call_agents_conversion | duration_agents_conversion) - all_agents
unmatched_retention = (call_agents_retention | duration_agents_retention) - all_agents

unmatched_sources = {
    'Conversion Agents Call Results': call_agents_conversion - all_agents,
    'Conversion Agents Duration Results': duration_agents_conversion - all_agents,
    'Retention Agents Call Results': call_agents_retention - all_agents,
    'Retention Agents Duration Results': duration_agents_retention - all_agents
}

print("\nUnmatched Agents:")
for source, unmatched in unmatched_sources.items():
    if unmatched:
        print(f"\n{source}:")
        for agent in unmatched:
            print(f" - {agent}")

print("\nAgent_Results.xlsx has been generated.")
