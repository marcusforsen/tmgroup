import pandas as pd
import numpy as np
import re

# Define file paths
call_results_path = r"C:\Users\marcus.forsen\Desktop\new project\Agent_Call_Results.xlsx"
duration_results_path = r"C:\Users\marcus.forsen\Desktop\new project\Agent_Duration_Results.xlsx"
output_path = r"C:\Users\marcus.forsen\Desktop\new project\Agent_Results.xlsx"

def safe_numeric_convert(series):
    """Convert a series to numeric, forcing errors to NaN."""
    return pd.to_numeric(series, errors='coerce')

def parse_duration(duration_str):
    """Convert a duration string into hh:mm:ss format."""
    if pd.isna(duration_str):
        return np.nan
    # Example: '2 h 30 m 15 s'
    hours = minutes = seconds = 0
    matches = re.findall(r'(\d+)\s*(h|m|s)', duration_str.lower())
    for value, unit in matches:
        if unit == 'h':
            hours = int(value)
        elif unit == 'm':
            minutes = int(value)
        elif unit == 's':
            seconds = int(value)
    return f"{hours:02}:{minutes:02}:{seconds:02}"

def parse_percentage(percentage_str):
    """Convert a percentage string to a float."""
    if pd.isna(percentage_str):
        return np.nan
    return float(percentage_str.replace('%', '').replace(',', '.'))

def find_column(df, possible_names):
    """Find the column name in DataFrame that matches any of the possible names."""
    for name in possible_names:
        if name in df.columns:
            return name
    return None

def process_sheet(call_sheet, duration_sheet, group):
    print(f"\nProcessing {group} sheet")
    print("Call sheet columns:", call_sheet.columns)
    print("Duration sheet columns:", duration_sheet.columns)
    
    # Parse duration strings to hh:mm:ss format
    duration_sheet['Total Time'] = duration_sheet['Total Time'].apply(parse_duration)
    
    # Convert target percentages to numeric
    call_sheet['Target Percentage'] = call_sheet['Target Percentage'].apply(parse_percentage)
    duration_sheet['Target Percentage'] = duration_sheet['Target Percentage'].apply(parse_percentage)
    
    # Merge dataframes
    merged_df = pd.merge(call_sheet, duration_sheet, on=['Agent Name', 'Desk'], suffixes=('_call', '_duration'))
    print("Merged DataFrame shape:", merged_df.shape)
    print("Merged DataFrame columns:", merged_df.columns)
    
    # Find column names
    call_attempts_col = find_column(merged_df, ['Call Attempts', 'Calls'])
    target_call_col = find_column(merged_df, ['Target_call', 'Target Call', 'Target'])
    duration_col = find_column(merged_df, ['Total Time'])  # From duration sheet
    target_duration_col = find_column(duration_sheet, ['Target'])  # From duration sheet
    target_percentage_call_col = find_column(merged_df, ['Target Percentage_call'])
    target_percentage_duration_col = find_column(merged_df, ['Target Percentage_duration'])
    
    print(f"Found columns: Call Attempts: {call_attempts_col}, Target Call: {target_call_col}, "
          f"Duration: {duration_col}, Target Duration: {target_duration_col}, "
          f"Target Percentage Call: {target_percentage_call_col}, Target Percentage Duration: {target_percentage_duration_col}")
    
    # Create new dataframe with required columns
    result_df = pd.DataFrame({
        'Group': group,
        'Agent Name': merged_df['Agent Name'],
        'Desk': merged_df['Desk'],
        'Call Attempts': safe_numeric_convert(merged_df[call_attempts_col]) if call_attempts_col else None,
        'Target (Call Attempts)': safe_numeric_convert(merged_df[target_call_col]) if target_call_col else None,
        'Duration': merged_df[duration_col] if duration_col else None,
        'Target (Duration)': duration_sheet[target_duration_col] if target_duration_col else None,
    })
    
    # Calculate Target Percentage Total
    if target_percentage_call_col and target_percentage_duration_col:
        target_percentage_call = safe_numeric_convert(merged_df[target_percentage_call_col])
        target_percentage_duration = safe_numeric_convert(merged_df[target_percentage_duration_col])
        result_df['Target Total'] = ((target_percentage_call + target_percentage_duration) / 2).apply(lambda x: f"{x:.2f}%")
    elif target_percentage_call_col:
        result_df['Target Total'] = safe_numeric_convert(merged_df[target_percentage_call_col]).apply(lambda x: f"{x:.2f}%")
    elif target_percentage_duration_col:
        result_df['Target Total'] = safe_numeric_convert(merged_df[target_percentage_duration_col]).apply(lambda x: f"{x:.2f}%")
    else:
        print(f"Warning: No 'Target Percentage' columns found in {group} sheet. Target Total not calculated.")
    
    print("Result DataFrame columns:", result_df.columns)
    print("Result DataFrame first few rows:")
    print(result_df.head())
    
    return result_df


# Read Excel files
call_df_conversion = pd.read_excel(call_results_path, sheet_name='Conversion Agents')
call_df_retention = pd.read_excel(call_results_path, sheet_name='Retention Agents')
duration_df_conversion = pd.read_excel(duration_results_path, sheet_name='Conversion Agents')
duration_df_retention = pd.read_excel(duration_results_path, sheet_name='Retention Agents')

# Process each sheet
result_conversion = process_sheet(call_df_conversion, duration_df_conversion, 'Conversion')
result_retention = process_sheet(call_df_retention, duration_df_retention, 'Retention')

# Combine results
final_result = pd.concat([result_conversion, result_retention], ignore_index=True)

# Define column order based on available columns
base_columns = ["Group", "Agent Name", "Desk", "Call Attempts", "Target (Call Attempts)", "Duration", "Target (Duration)"]
if 'Target Total' in final_result.columns:
    base_columns.append('Target Total')

# Reorder columns
final_result = final_result[base_columns]

# Write to Excel
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    final_result.to_excel(writer, index=False, sheet_name='Combined Results')
    result_conversion.to_excel(writer, index=False, sheet_name='Conversion Agents')
    result_retention.to_excel(writer, index=False, sheet_name='Retention Agents')

print(f"\nProcessing complete. Output file saved at: {output_path}")
print("Columns in final_result:", final_result.columns)
print("Number of rows in final_result:", len(final_result))

# Print detailed information about input files
for df_name, df in [("call_df_conversion", call_df_conversion), 
                    ("duration_df_conversion", duration_df_conversion),
                    ("call_df_retention", call_df_retention),
                    ("duration_df_retention", duration_df_retention)]:
    print(f"\nDetailed information for {df_name}:")
    print(f"Shape: {df.shape}")
    print("Columns:")
    for col in df.columns:
        non_null_count = df[col].count()
        first_value = df[col].dropna().iloc[0] if not df[col].dropna().empty else "No non-null values"
        print(f"  {col}: {df[col].dtype}, Non-null count: {non_null_count}, First non-null value: {first_value}")
    print("\nFirst few rows:")
    print(df.head())
    print("\n" + "="*50 + "\n")




























import pandas as pd
import re

def process_sheet(call_sheet, duration_sheet, group):
    print(f"\nProcessing {group} sheet")
    print("Call sheet columns:", call_sheet.columns)
    print("Duration sheet columns:", duration_sheet.columns)
    
    # Parse duration strings to hh:mm:ss format
    duration_sheet['Total Time'] = duration_sheet['Total Time'].apply(parse_duration)
    
    # Convert target percentages to numeric
    call_sheet['Target Percentage'] = call_sheet['Target Percentage'].apply(parse_percentage)
    duration_sheet['Target Percentage'] = duration_sheet['Target Percentage'].apply(parse_percentage)
    
    # Merge dataframes
    merged_df = pd.merge(call_sheet, duration_sheet, on=['Agent Name', 'Desk'], suffixes=('_call', '_duration'))
    print("Merged DataFrame shape:", merged_df.shape)
    print("Merged DataFrame columns:", merged_df.columns)
    
    # Find column names
    call_attempts_col = find_column(merged_df, ['Call Attempts', 'Calls'])
    target_call_col = find_column(merged_df, ['Target_call', 'Target Call', 'Target'])
    duration_col = find_column(merged_df, ['Total Time'])  # From duration sheet
    target_duration_col = find_column(merged_df, ['Target_duration'])  # From merged DataFrame
    target_percentage_call_col = find_column(merged_df, ['Target Percentage_call'])
    target_percentage_duration_col = find_column(merged_df, ['Target Percentage_duration'])
    
    print(f"Found columns: Call Attempts: {call_attempts_col}, Target Call: {target_call_col}, "
          f"Duration: {duration_col}, Target Duration: {target_duration_col}, "
          f"Target Percentage Call: {target_percentage_call_col}, Target Percentage Duration: {target_percentage_duration_col}")
    
    # Create new dataframe with required columns
    result_df = pd.DataFrame({
        'Group': group,
        'Agent Name': merged_df['Agent Name'],
        'Desk': merged_df['Desk'],
        'Call Attempts': safe_numeric_convert(merged_df[call_attempts_col]) if call_attempts_col else None,
        'Target (Call Attempts)': safe_numeric_convert(merged_df[target_call_col]) if target_call_col else None,
        'Target % (Call Attempts)': merged_df[target_percentage_call_col].apply(lambda x: f"{x * 100:.2f}%") if target_percentage_call_col else None,
        'Duration': merged_df[duration_col] if duration_col else None,
        'Target (Duration)': merged_df[target_duration_col] if target_duration_col else None,
        'Target % (Duration)': merged_df[target_percentage_duration_col].apply(lambda x: f"{x * 100:.2f}%") if target_percentage_duration_col else None,
    })
    
    # Calculate Target Percentage Total
    def calculate_target_total(row):
        call_percentage = row.get('Target % (Call Attempts)', '0%').strip('%')
        duration_percentage = row.get('Target % (Duration)', '0%').strip('%')
        
        try:
            call_percentage = float(call_percentage) if call_percentage else 0
            duration_percentage = float(duration_percentage) if duration_percentage else 0
            target_total = (call_percentage + duration_percentage) / 2
            return f"{target_total:.2f}%"
        except ValueError:
            return 'NaN'
    
    result_df['Target Total'] = result_df.apply(calculate_target_total, axis=1)
    
    print("Result DataFrame columns:", result_df.columns)
    print("Result DataFrame first few rows:")
    print(result_df.head())
    
    return result_df

def find_column(df, possible_names):
    for name in possible_names:
        if name in df.columns:
            return name
    print(f"Warning: None of {possible_names} found in columns.")
    return None

def parse_percentage(percentage_str):
    if pd.isna(percentage_str):
        return None
    # Ensure percentage is in decimal format
    return float(percentage_str.strip('%')) / 100.0 if '%' in percentage_str else float(percentage_str)

def parse_duration(duration_str):
    # Convert to hh:mm:ss format
    if pd.isna(duration_str):
        return None
    
    # Extract time components using regex
    hours = minutes = seconds = 0
    time_parts = re.findall(r'(\d+)\s*(h|m|s)', duration_str.lower())
    
    for value, unit in time_parts:
        if unit == 'h':
            hours = int(value)
        elif unit == 'm':
            minutes = int(value)
        elif unit == 's':
            seconds = int(value)
    
    # Format as hh:mm:ss
    return f"{hours:02}:{minutes:02}:{seconds:02}"

def safe_numeric_convert(series):
    return pd.to_numeric(series, errors='coerce')

# Load data from files for both sheets
call_df_conversion = pd.read_excel('Agent_Call_Results.xlsx', sheet_name='Conversion Agents')
duration_df_conversion = pd.read_excel('Agent_Duration_Results.xlsx', sheet_name='Conversion Agents')

# Process Conversion agents
result_conversion = process_sheet(call_df_conversion, duration_df_conversion, 'Conversion')

# Load data from files for Retention agents
call_df_retention = pd.read_excel('Agent_Call_Results.xlsx', sheet_name='Retention Agents')
duration_df_retention = pd.read_excel('Agent_Duration_Results.xlsx', sheet_name='Retention Agents')

# Process Retention agents
result_retention = process_sheet(call_df_retention, duration_df_retention, 'Retention')

# Save results to separate sheets in the same Excel file
with pd.ExcelWriter('Agent_Results.xlsx') as writer:
    result_conversion.to_excel(writer, sheet_name='Conversion Agents', index=False)
    result_retention.to_excel(writer, sheet_name='Retention Agents', index=False)



























import pandas as pd
import re

def extract_sources(sources_str, suffix):
    sources_dict = {}
    if pd.isna(sources_str):
        return sources_dict
    
    print(f"Parsing sources: {sources_str}")  # Debug print

    # Split the string into individual source entries
    entries = re.split(r',\s*', sources_str)
    
    # Define the sources of interest
    new_columns = ['Voiso Traling', 'Voiso Summitlife', 'Coperato Traling', 'Coperato Signix', 'Voicespin']

    for entry in entries:
        match = re.match(r'([a-zA-Z\s]+)\.csv:\s*([\d\s\w]*)', entry)
        if match:
            source_name = match.group(1).strip()
            source_value = match.group(2).strip()
            
            # Clean source_name and check if it is in new_columns
            for column in new_columns:
                if source_name.lower() in column.lower():
                    if suffix == "(minutes)":
                        # Parse duration string to minutes
                        source_value = parse_duration(source_value)
                    else:
                        # Convert to integer for calls
                        source_value = int(source_value)
                    
                    sources_dict[column + f" {suffix}"] = source_value
                    break
    
    print(f"Extracted sources: {sources_dict}")  # Debug print
    return sources_dict

def parse_duration(duration_str):
    if pd.isna(duration_str):
        return 0
    
    total_seconds = 0
    # Match hours, minutes, and seconds
    hours_match = re.search(r'(\d+)\s*h', duration_str)
    minutes_match = re.search(r'(\d+)\s*m', duration_str)
    seconds_match = re.search(r'(\d+)\s*s', duration_str)
    
    if hours_match:
        total_seconds += int(hours_match.group(1)) * 3600
    if minutes_match:
        total_seconds += int(minutes_match.group(1)) * 60
    if seconds_match:
        total_seconds += int(seconds_match.group(1))
    
    # Convert seconds to minutes
    total_minutes = total_seconds / 60
    return round(total_minutes, 2)  # Return minutes rounded to 2 decimal places

def parse_percentage(percentage_str):
    try:
        return float(percentage_str.strip('%')) / 100
    except ValueError:
        return 0.0

def find_column(df, names):
    for name in names:
        if name in df.columns:
            return name
    return None

def safe_numeric_convert(series):
    return pd.to_numeric(series, errors='coerce')

def process_sheet(call_sheet, duration_sheet, group):
    print(f"\nProcessing {group} sheet")
    print("Call sheet columns:", call_sheet.columns)
    print("Duration sheet columns:", duration_sheet.columns)
    
    # Convert target percentages to numeric
    call_sheet['Target Percentage'] = call_sheet['Target Percentage'].apply(parse_percentage)
    duration_sheet['Target Percentage'] = duration_sheet['Target Percentage'].apply(parse_percentage)
    
    # Merge dataframes
    merged_df = pd.merge(call_sheet, duration_sheet, on=['Agent Name', 'Desk'], suffixes=('_call', '_duration'))
    print("Merged DataFrame shape:", merged_df.shape)
    print("Merged DataFrame columns:", merged_df.columns)
    
    # Find column names
    call_attempts_col = find_column(merged_df, ['Call Attempts', 'Calls'])
    target_call_col = find_column(merged_df, ['Target_call', 'Target Call', 'Target'])
    duration_col = find_column(merged_df, ['Total Time'])  # From duration sheet
    target_duration_col = find_column(merged_df, ['Target_duration'])  # From merged DataFrame
    target_percentage_call_col = find_column(merged_df, ['Target Percentage_call'])
    target_percentage_duration_col = find_column(merged_df, ['Target Percentage_duration'])
    sources_call_col = find_column(merged_df, ['Sources_call'])
    sources_duration_col = find_column(merged_df, ['Sources_duration'])
    
    print(f"Found columns: Call Attempts: {call_attempts_col}, Target Call: {target_call_col}, "
          f"Duration: {duration_col}, Target Duration: {target_duration_col}, "
          f"Target Percentage Call: {target_percentage_call_col}, Target Percentage Duration: {target_percentage_duration_col}, "
          f"Sources Call: {sources_call_col}, Sources Duration: {sources_duration_col}")
    
    # Create new dataframe with required columns
    result_df = pd.DataFrame({
        'Group': group,
        'Agent Name': merged_df['Agent Name'],
        'Desk': merged_df['Desk'],
        'Call Attempts': safe_numeric_convert(merged_df[call_attempts_col]) if call_attempts_col else None,
        'Target (Call Attempts)': safe_numeric_convert(merged_df[target_call_col]) if target_call_col else None,
        'Target % (Call Attempts)': merged_df[target_percentage_call_col].apply(lambda x: f"{x * 100:.2f}%") if target_percentage_call_col else None,
        'Duration': merged_df[duration_col] if duration_col else None,
        'Target (Duration)': merged_df[target_duration_col] if target_duration_col else None,
        'Target % (Duration)': merged_df[target_percentage_duration_col].apply(lambda x: f"{x * 100:.2f}%") if target_percentage_duration_col else None,
    })
    
    # Add new columns for sources with correct suffixes
    new_columns = ['Voiso Traling', 'Voiso Summitlife', 'Coperato Traling', 'Coperato Signix', 'Voicespin']
    for col in new_columns:
        result_df[col + " (Calls)"] = 0
        result_df[col + " (minutes)"] = 0

    # Extract sources and fill new columns
    for index, row in merged_df.iterrows():
        sources_call = row[sources_call_col] if sources_call_col else ""
        sources_duration = row[sources_duration_col] if sources_duration_col else ""
        
        # Add suffixes for calls and minutes
        sources_dict_call = extract_sources(sources_call, "(Calls)")
        sources_dict_duration = extract_sources(sources_duration, "(minutes)")
        
        for col in new_columns:
            result_df.at[index, col + " (Calls)"] = sources_dict_call.get(col + " (Calls)", 0)
            result_df.at[index, col + " (minutes)"] = sources_dict_duration.get(col + " (minutes)", 0)
    
    # Calculate Target Percentage Total
    def calculate_target_total(row):
        call_percentage = row.get('Target % (Call Attempts)', '0%').strip('%')
        duration_percentage = row.get('Target % (Duration)', '0%').strip('%')
        
        try:
            call_percentage = float(call_percentage) if call_percentage else 0
            duration_percentage = float(duration_percentage) if duration_percentage else 0
            target_total = (call_percentage + duration_percentage) / 2
            return f"{target_total:.2f}%"
        except ValueError:
            return 'NaN'
    
    result_df['Target Total'] = result_df.apply(calculate_target_total, axis=1)
    
    print("Result DataFrame columns:", result_df.columns)
    print("Result DataFrame first few rows:")
    print(result_df.head())
    
    return result_df

# Load data from files for both sheets
call_df_conversion = pd.read_excel('Agent_Call_Results.xlsx', sheet_name='Conversion Agents')
duration_df_conversion = pd.read_excel('Agent_Duration_Results.xlsx', sheet_name='Conversion Agents')

# Process Conversion agents
result_conversion = process_sheet(call_df_conversion, duration_df_conversion, 'Conversion')

# Load data from files for Retention agents
call_df_retention = pd.read_excel('Agent_Call_Results.xlsx', sheet_name='Retention Agents')
duration_df_retention = pd.read_excel('Agent_Duration_Results.xlsx', sheet_name='Retention Agents')

# Process Retention agents
result_retention = process_sheet(call_df_retention, duration_df_retention, 'Retention')

# Save results to separate sheets in the same Excel file
with pd.ExcelWriter('Agent_Results.xlsx') as writer:
    result_conversion.to_excel(writer, sheet_name='Conversion Agents', index=False)
    result_retention.to_excel(writer, sheet_name='Retention Agents', index=False)








































import pandas as pd
import re

def extract_sources(sources_str, suffix):
    sources_dict = {}
    if pd.isna(sources_str):
        return sources_dict
    
    print(f"Parsing sources: {sources_str}")  # Debug print

    # Split the string into individual source entries
    entries = re.split(r',\s*', sources_str)
    
    # Define the sources of interest
    new_columns = ['Voiso Traling', 'Voiso Summitlife', 'Coperato Traling', 'Coperato Signix', 'Voicespin']

    for entry in entries:
        match = re.match(r'([a-zA-Z\s]+)\.csv:\s*([\d\s\w]*)', entry)
        if match:
            source_name = match.group(1).strip()
            source_value = match.group(2).strip()
            
            # Clean source_name and check if it is in new_columns
            for column in new_columns:
                if source_name.lower() in column.lower():
                    if suffix == "(minutes)":
                        # Parse duration string to formatted hours and minutes
                        source_value = parse_duration(source_value)
                    else:
                        # Convert to integer for calls
                        source_value = int(source_value)
                    
                    sources_dict[column + f" {suffix}"] = source_value
                    break
    
    print(f"Extracted sources: {sources_dict}")  # Debug print
    return sources_dict

def parse_duration(duration_str):
    if pd.isna(duration_str):
        return "0m"
    
    total_seconds = 0
    # Match hours, minutes, and seconds
    hours_match = re.search(r'(\d+)\s*h', duration_str)
    minutes_match = re.search(r'(\d+)\s*m', duration_str)
    seconds_match = re.search(r'(\d+)\s*s', duration_str)
    
    if hours_match:
        total_seconds += int(hours_match.group(1)) * 3600
    if minutes_match:
        total_seconds += int(minutes_match.group(1)) * 60
    if seconds_match:
        total_seconds += int(seconds_match.group(1))
    
    # Convert total seconds to hours and minutes
    total_minutes = total_seconds // 60
    hours = total_minutes // 60
    minutes = total_minutes % 60
    
    if hours > 0:
        return f"{hours}h {minutes}m"
    else:
        return f"{minutes}m"

def parse_percentage(percentage_str):
    try:
        return float(percentage_str.strip('%')) / 100
    except ValueError:
        return 0.0

def find_column(df, names):
    for name in names:
        if name in df.columns:
            return name
    return None

def safe_numeric_convert(series):
    return pd.to_numeric(series, errors='coerce')

def process_sheet(call_sheet, duration_sheet, group):
    print(f"\nProcessing {group} sheet")
    print("Call sheet columns:", call_sheet.columns)
    print("Duration sheet columns:", duration_sheet.columns)
    
    # Convert target percentages to numeric
    call_sheet['Target Percentage'] = call_sheet['Target Percentage'].apply(parse_percentage)
    duration_sheet['Target Percentage'] = duration_sheet['Target Percentage'].apply(parse_percentage)
    
    # Merge dataframes
    merged_df = pd.merge(call_sheet, duration_sheet, on=['Agent Name', 'Desk'], suffixes=('_call', '_duration'))
    print("Merged DataFrame shape:", merged_df.shape)
    print("Merged DataFrame columns:", merged_df.columns)
    
    # Find column names
    call_attempts_col = find_column(merged_df, ['Call Attempts', 'Calls'])
    target_call_col = find_column(merged_df, ['Target_call', 'Target Call', 'Target'])
    duration_col = find_column(merged_df, ['Total Time'])  # From duration sheet
    target_duration_col = find_column(merged_df, ['Target_duration'])  # From merged DataFrame
    target_percentage_call_col = find_column(merged_df, ['Target Percentage_call'])
    target_percentage_duration_col = find_column(merged_df, ['Target Percentage_duration'])
    sources_call_col = find_column(merged_df, ['Sources_call'])
    sources_duration_col = find_column(merged_df, ['Sources_duration'])
    
    print(f"Found columns: Call Attempts: {call_attempts_col}, Target Call: {target_call_col}, "
          f"Duration: {duration_col}, Target Duration: {target_duration_col}, "
          f"Target Percentage Call: {target_percentage_call_col}, Target Percentage Duration: {target_percentage_duration_col}, "
          f"Sources Call: {sources_call_col}, Sources Duration: {sources_duration_col}")
    
    # Create new dataframe with required columns
    result_df = pd.DataFrame({
        'Group': group,
        'Agent Name': merged_df['Agent Name'],
        'Desk': merged_df['Desk'],
        'Call Attempts': safe_numeric_convert(merged_df[call_attempts_col]) if call_attempts_col else None,
        'Target (Call Attempts)': safe_numeric_convert(merged_df[target_call_col]) if target_call_col else None,
        'Target % (Call Attempts)': merged_df[target_percentage_call_col].apply(lambda x: f"{x * 100:.2f}%") if target_percentage_call_col else None,
        'Duration': merged_df[duration_col] if duration_col else None,
        'Target (Duration)': merged_df[target_duration_col] if target_duration_col else None,
        'Target % (Duration)': merged_df[target_percentage_duration_col].apply(lambda x: f"{x * 100:.2f}%") if target_percentage_duration_col else None,
    })
    
    # Add new columns for sources with correct suffixes
    new_columns = ['Voiso Traling', 'Voiso Summitlife', 'Coperato Traling', 'Coperato Signix', 'Voicespin']
    for col in new_columns:
        result_df[col + " (Calls)"] = 0
        result_df[col + " (minutes)"] = "0m"

    # Extract sources and fill new columns
    for index, row in merged_df.iterrows():
        sources_call = row[sources_call_col] if sources_call_col else ""
        sources_duration = row[sources_duration_col] if sources_duration_col else ""
        
        # Add suffixes for calls and minutes
        sources_dict_call = extract_sources(sources_call, "(Calls)")
        sources_dict_duration = extract_sources(sources_duration, "(minutes)")
        
        for col in new_columns:
            result_df.at[index, col + " (Calls)"] = sources_dict_call.get(col + " (Calls)", 0)
            result_df.at[index, col + " (minutes)"] = sources_dict_duration.get(col + " (minutes)", "0m")
    
    # Calculate Target Percentage Total
    def calculate_target_total(row):
        call_percentage = row.get('Target % (Call Attempts)', '0%').strip('%')
        duration_percentage = row.get('Target % (Duration)', '0%').strip('%')
        
        try:
            call_percentage = float(call_percentage) if call_percentage else 0
            duration_percentage = float(duration_percentage) if duration_percentage else 0
            target_total = (call_percentage + duration_percentage) / 2
            return f"{target_total:.2f}%"
        except ValueError:
            return 'NaN'
    
    result_df['Target Total'] = result_df.apply(calculate_target_total, axis=1)
    
    print("Result DataFrame columns:", result_df.columns)
    print("Result DataFrame first few rows:")
    print(result_df.head())
    
    return result_df

# Load data from files for both sheets
call_df_conversion = pd.read_excel('Agent_Call_Results.xlsx', sheet_name='Conversion Agents')
duration_df_conversion = pd.read_excel('Agent_Duration_Results.xlsx', sheet_name='Conversion Agents')

# Process Conversion agents
result_conversion = process_sheet(call_df_conversion, duration_df_conversion, 'Conversion')

# Load data from files for Retention agents
call_df_retention = pd.read_excel('Agent_Call_Results.xlsx', sheet_name='Retention Agents')
duration_df_retention = pd.read_excel('Agent_Duration_Results.xlsx', sheet_name='Retention Agents')

# Process Retention agents
result_retention = process_sheet(call_df_retention, duration_df_retention, 'Retention')

# Save results to separate sheets in the same Excel file
with pd.ExcelWriter('Agent_Results.xlsx') as writer:
    result_conversion.to_excel(writer, sheet_name='Conversion Agents', index=False)
    result_retention.to_excel(writer, sheet_name='Retention Agents', index=False)









































import pandas as pd
import re

def extract_sources(sources_str, suffix):
    sources_dict = {}
    if pd.isna(sources_str):
        return sources_dict
    
    print(f"Parsing sources: {sources_str}")  # Debug print

    # Split the string into individual source entries
    entries = re.split(r',\s*', sources_str)
    
    # Define the sources of interest
    new_columns = ['Voiso Traling', 'Voiso Summitlife', 'Coperato Traling', 'Coperato Signix', 'Voicespin']

    for entry in entries:
        match = re.match(r'([a-zA-Z\s]+)\.csv:\s*([\d\s\w]*)', entry)
        if match:
            source_name = match.group(1).strip()
            source_value = match.group(2).strip()
            
            # Clean source_name and check if it is in new_columns
            for column in new_columns:
                if source_name.lower() in column.lower():
                    if suffix == "(Calls)":
                        # Convert to integer for calls
                        source_value = int(source_value)
                    
                    sources_dict[column + f" {suffix}"] = source_value
                    break
    
    print(f"Extracted sources: {sources_dict}")  # Debug print
    return sources_dict

def parse_percentage(percentage_str):
    try:
        return float(percentage_str.strip('%')) / 100
    except ValueError:
        return 0.0

def find_column(df, names):
    for name in names:
        if name in df.columns:
            return name
    return None

def safe_numeric_convert(series):
    return pd.to_numeric(series, errors='coerce')

def process_sheet(call_sheet, duration_sheet, group):
    print(f"\nProcessing {group} sheet")
    print("Call sheet columns:", call_sheet.columns)
    print("Duration sheet columns:", duration_sheet.columns)
    
    # Convert target percentages to numeric
    call_sheet['Target Percentage'] = call_sheet['Target Percentage'].apply(parse_percentage)
    duration_sheet['Target Percentage'] = duration_sheet['Target Percentage'].apply(parse_percentage)
    
    # Merge dataframes
    merged_df = pd.merge(call_sheet, duration_sheet, on=['Agent Name', 'Desk'], suffixes=('_call', '_duration'))
    print("Merged DataFrame shape:", merged_df.shape)
    print("Merged DataFrame columns:", merged_df.columns)
    
    # Find column names
    call_attempts_col = find_column(merged_df, ['Call Attempts', 'Calls'])
    target_call_col = find_column(merged_df, ['Target_call', 'Target Call', 'Target'])
    duration_col = find_column(merged_df, ['Total Time'])  # From duration sheet
    target_duration_col = find_column(merged_df, ['Target_duration'])  # From merged DataFrame
    target_percentage_call_col = find_column(merged_df, ['Target Percentage_call'])
    target_percentage_duration_col = find_column(merged_df, ['Target Percentage_duration'])
    sources_call_col = find_column(merged_df, ['Sources_call'])
    
    print(f"Found columns: Call Attempts: {call_attempts_col}, Target Call: {target_call_col}, "
          f"Duration: {duration_col}, Target Duration: {target_duration_col}, "
          f"Target Percentage Call: {target_percentage_call_col}, Target Percentage Duration: {target_percentage_duration_col}, "
          f"Sources Call: {sources_call_col}")
    
    # Create new dataframe with required columns
    result_df = pd.DataFrame({
        'Group': group,
        'Agent Name': merged_df['Agent Name'],
        'Desk': merged_df['Desk'],
        'Call Attempts': safe_numeric_convert(merged_df[call_attempts_col]) if call_attempts_col else None,
        'Target (Call Attempts)': safe_numeric_convert(merged_df[target_call_col]) if target_call_col else None,
        'Target % (Call Attempts)': merged_df[target_percentage_call_col].apply(lambda x: f"{x * 100:.2f}%") if target_percentage_call_col else None,
        'Duration': merged_df[duration_col] if duration_col else None,
        'Target (Duration)': merged_df[target_duration_col] if target_duration_col else None,
        'Target % (Duration)': merged_df[target_percentage_duration_col].apply(lambda x: f"{x * 100:.2f}%") if target_percentage_duration_col else None,
    })
    
    # Add new columns for sources with correct suffixes
    new_columns = ['Voiso Traling', 'Voiso Summitlife', 'Coperato Traling', 'Coperato Signix', 'Voicespin']
    for col in new_columns:
        result_df[col + " % Calls"] = 0.0  # Percentage column initialized to 0.0

    # Extract sources and fill new columns
    for index, row in merged_df.iterrows():
        sources_call = row[sources_call_col] if sources_call_col else ""
        
        # Add suffix for calls
        sources_dict_call = extract_sources(sources_call, "(Calls)")
        
        total_calls = row[call_attempts_col] if call_attempts_col else 0
        for col in new_columns:
            calls = sources_dict_call.get(col + " (Calls)", 0)
            if total_calls > 0:
                result_df.at[index, col + " % Calls"] = (calls / total_calls) * 100

    # Calculate Target Percentage Total
    def calculate_target_total(row):
        call_percentage = row.get('Target % (Call Attempts)', '0%').strip('%')
        duration_percentage = row.get('Target % (Duration)', '0%').strip('%')
        
        try:
            call_percentage = float(call_percentage) if call_percentage else 0
            duration_percentage = float(duration_percentage) if duration_percentage else 0
            target_total = (call_percentage + duration_percentage) / 2
            return f"{target_total:.2f}%"
        except ValueError:
            return 'NaN'
    
    result_df['Target Total'] = result_df.apply(calculate_target_total, axis=1)
    
    # Remove columns with "(Calls)" in their names
    result_df = result_df.loc[:, ~result_df.columns.str.contains(r'\(Calls\)', case=False)]

    # Add source % Calls columns back to the result_df
    for col in new_columns:
        if col + " % Calls" in result_df.columns:
            result_df[col + " % Calls"] = result_df[col + " % Calls"].apply(lambda x: f"{x:.2f}%")
    
    print("Result DataFrame columns:", result_df.columns)
    print("Result DataFrame first few rows:")
    print(result_df.head())
    
    return result_df

# Load data from files for both sheets
call_df_conversion = pd.read_excel('Agent_Call_Results.xlsx', sheet_name='Conversion Agents')
duration_df_conversion = pd.read_excel('Agent_Duration_Results.xlsx', sheet_name='Conversion Agents')

# Process Conversion agents
result_conversion = process_sheet(call_df_conversion, duration_df_conversion, 'Conversion')

# Load data from files for Retention agents
call_df_retention = pd.read_excel('Agent_Call_Results.xlsx', sheet_name='Retention Agents')
duration_df_retention = pd.read_excel('Agent_Duration_Results.xlsx', sheet_name='Retention Agents')

# Process Retention agents
result_retention = process_sheet(call_df_retention, duration_df_retention, 'Retention')

# Save results to separate sheets in the same Excel file
with pd.ExcelWriter('Agent_Results.xlsx') as writer:
    result_conversion.to_excel(writer, sheet_name='Conversion Agents', index=False)
    result_retention.to_excel(writer, sheet_name='Retention Agents', index=False)


























































import pandas as pd
import re

def extract_sources(sources_str, suffix):
    sources_dict = {}
    if pd.isna(sources_str):
        return sources_dict
    
    print(f"Parsing sources: {sources_str}")  # Debug print

    # Split the string into individual source entries
    entries = re.split(r',\s*', sources_str)
    
    # Define the sources of interest
    new_columns = ['Voiso Traling', 'Voiso Summitlife', 'Coperato Traling', 'Coperato Signix', 'Voicespin']

    for entry in entries:
        match = re.match(r'([a-zA-Z\s]+)\.csv:\s*([\d\s\w]*)', entry)
        if match:
            source_name = match.group(1).strip()
            source_value = match.group(2).strip()
            
            # Clean source_name and check if it is in new_columns
            for column in new_columns:
                if source_name.lower() in column.lower():
                    if suffix == "(Calls)":
                        # Convert to integer for calls
                        source_value = int(source_value)
                    
                    sources_dict[column + f" {suffix}"] = source_value
                    break
    
    print(f"Extracted sources: {sources_dict}")  # Debug print
    return sources_dict

def parse_percentage(percentage_str):
    try:
        return float(percentage_str.strip('%')) / 100
    except ValueError:
        return 0.0

def find_column(df, names):
    for name in names:
        if name in df.columns:
            return name
    return None

def safe_numeric_convert(series):
    return pd.to_numeric(series, errors='coerce')

def process_sheet(call_sheet, duration_sheet, group):
    print(f"\nProcessing {group} sheet")
    print("Call sheet columns:", call_sheet.columns)
    print("Duration sheet columns:", duration_sheet.columns)
    
    # Convert target percentages to numeric
    call_sheet['Target Percentage'] = call_sheet['Target Percentage'].apply(parse_percentage)
    duration_sheet['Target Percentage'] = duration_sheet['Target Percentage'].apply(parse_percentage)
    
    # Merge dataframes
    merged_df = pd.merge(call_sheet, duration_sheet, on=['Agent Name', 'Desk'], suffixes=('_call', '_duration'))
    print("Merged DataFrame shape:", merged_df.shape)
    print("Merged DataFrame columns:", merged_df.columns)
    
    # Find column names
    call_attempts_col = find_column(merged_df, ['Call Attempts', 'Calls'])
    target_call_col = find_column(merged_df, ['Target_call', 'Target Call', 'Target'])
    duration_col = find_column(merged_df, ['Total Time'])  # From duration sheet
    target_duration_col = find_column(merged_df, ['Target_duration'])  # From merged DataFrame
    target_percentage_call_col = find_column(merged_df, ['Target Percentage_call'])
    target_percentage_duration_col = find_column(merged_df, ['Target Percentage_duration'])
    sources_call_col = find_column(merged_df, ['Sources_call'])
    
    print(f"Found columns: Call Attempts: {call_attempts_col}, Target Call: {target_call_col}, "
          f"Duration: {duration_col}, Target Duration: {target_duration_col}, "
          f"Target Percentage Call: {target_percentage_call_col}, Target Percentage Duration: {target_percentage_duration_col}, "
          f"Sources Call: {sources_call_col}")
    
    # Create new dataframe with required columns
    result_df = pd.DataFrame({
        'Group': group,
        'Agent Name': merged_df['Agent Name'],
        'Desk': merged_df['Desk'],
        'Call Attempts': safe_numeric_convert(merged_df[call_attempts_col]) if call_attempts_col else None,
        'Target (Call Attempts)': safe_numeric_convert(merged_df[target_call_col]) if target_call_col else None,
        'Target % (Call Attempts)': merged_df[target_percentage_call_col].apply(lambda x: f"{x * 100:.2f}%") if target_percentage_call_col else None,
        'Duration': merged_df[duration_col] if duration_col else None,
        'Target (Duration)': merged_df[target_duration_col] if target_duration_col else None,
        'Target % (Duration)': merged_df[target_percentage_duration_col].apply(lambda x: f"{x * 100:.2f}%") if target_percentage_duration_col else None,
    })
    
    # Add new columns for sources with correct suffixes
    new_columns = ['Voiso Traling', 'Voiso Summitlife', 'Coperato Traling', 'Coperato Signix', 'Voicespin']
    for col in new_columns:
        result_df[col + " %"] = 0.0  # Percentage column initialized to 0.0

    # Extract sources and fill new columns
    for index, row in merged_df.iterrows():
        sources_call = row[sources_call_col] if sources_call_col else ""
        
        # Add suffix for calls
        sources_dict_call = extract_sources(sources_call, "(Calls)")
        
        total_calls = row[call_attempts_col] if call_attempts_col else 0
        for col in new_columns:
            calls = sources_dict_call.get(col + " (Calls)", 0)
            if total_calls > 0:
                result_df.at[index, col + " %"] = (calls / total_calls) * 100

    # Calculate Target Percentage Total
    def calculate_target_total(row):
        call_percentage = row.get('Target % (Call Attempts)', '0%').strip('%')
        duration_percentage = row.get('Target % (Duration)', '0%').strip('%')
        
        try:
            call_percentage = float(call_percentage) if call_percentage else 0
            duration_percentage = float(duration_percentage) if duration_percentage else 0
            target_total = (call_percentage + duration_percentage) / 2
            return f"{target_total:.2f}%"
        except ValueError:
            return 'NaN'
    
    result_df['Target Total'] = result_df.apply(calculate_target_total, axis=1)
    
    # Remove columns with "(Calls)" in their names
    result_df = result_df.loc[:, ~result_df.columns.str.contains(r'\(Calls\)', case=False)]

    # Rename columns by removing "(Calls)" suffix
    rename_columns = {col: col.replace(' (Calls)', '') for col in result_df.columns if ' (Calls)' in col}
    result_df.rename(columns=rename_columns, inplace=True)

    # Ensure percentage columns are formatted correctly
    for col in new_columns:
        if col + " %" in result_df.columns:
            result_df[col + " %"] = result_df[col + " %"].apply(lambda x: f"{x:.2f}%")
    
    print("Result DataFrame columns:", result_df.columns)
    print("Result DataFrame first few rows:")
    print(result_df.head())
    
    return result_df

# Load data from files for both sheets
call_df_conversion = pd.read_excel('Agent_Call_Results.xlsx', sheet_name='Conversion Agents')
duration_df_conversion = pd.read_excel('Agent_Duration_Results.xlsx', sheet_name='Conversion Agents')

# Process Conversion agents
result_conversion = process_sheet(call_df_conversion, duration_df_conversion, 'Conversion')

# Load data from files for Retention agents
call_df_retention = pd.read_excel('Agent_Call_Results.xlsx', sheet_name='Retention Agents')
duration_df_retention = pd.read_excel('Agent_Duration_Results.xlsx', sheet_name='Retention Agents')

# Process Retention agents
result_retention = process_sheet(call_df_retention, duration_df_retention, 'Retention')

# Save results to separate sheets in the same Excel file
with pd.ExcelWriter('Agent_Results.xlsx') as writer:
    result_conversion.to_excel(writer, sheet_name='Conversion Agents', index=False)
    result_retention.to_excel(writer, sheet_name='Retention Agents', index=False)










































import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side

# Load the data from the Excel files
agent_list = pd.read_excel(r"C:\Users\marcus.forsen\Desktop\new project\agents.xlsx")
call_results_conversion = pd.read_excel(r"C:\Users\marcus.forsen\Desktop\new project\Agent_Call_Results.xlsx", sheet_name='Conversion Agents')
call_results_retention = pd.read_excel(r"C:\Users\marcus.forsen\Desktop\new project\Agent_Call_Results.xlsx", sheet_name='Retention Agents')
duration_results_conversion = pd.read_excel(r"C:\Users\marcus.forsen\Desktop\new project\Agent_Duration_Results.xlsx", sheet_name='Conversion Agents')
duration_results_retention = pd.read_excel(r"C:\Users\marcus.forsen\Desktop\new project\Agent_Duration_Results.xlsx", sheet_name='Retention Agents')

# Clean and standardize the 'Agent Name' in all dataframes
agent_list['AGENTNAME'] = agent_list['AGENTNAME'].str.strip().str.upper()
call_results_conversion['Agent Name'] = call_results_conversion['Agent Name'].str.strip().str.upper()
call_results_retention['Agent Name'] = call_results_retention['Agent Name'].str.strip().str.upper()
duration_results_conversion['Agent Name'] = duration_results_conversion['Agent Name'].str.strip().str.upper()
duration_results_retention['Agent Name'] = duration_results_retention['Agent Name'].str.strip().str.upper()

# Extract and rename columns from the agent list
agents_data = agent_list[['AGENTNAME', 'DESK', 'DEPARTMENT']].copy()
agents_data.rename(columns={'AGENTNAME': 'Agent Name', 'DESK': 'Desk'}, inplace=True)

# Define the custom sorting order
desk_order = [
    'Japanese', 'Korea', 'India Ajay', 'India Aarav', 'India', 'English', 'French', 
    'German', 'Spanish', 'Portuguese'
]
desk_order_dict = {desk: idx for idx, desk in enumerate(desk_order)}

# Extract relevant columns and handle call results and durations for Conversion Agents
call_data_conversion = call_results_conversion[['Agent Name', 'Unique', 'Call Attempts']].copy()
duration_data_conversion = duration_results_conversion[['Agent Name', 'Total Time']].copy()

# Convert the "Total Time" to timedelta and handle formatting
duration_data_conversion['Total Time'] = pd.to_timedelta(duration_data_conversion['Total Time'], errors='coerce')
duration_data_conversion['Duration'] = duration_data_conversion['Total Time'].apply(lambda x: str(x).split()[-1] if pd.notnull(x) else '0')
duration_data_conversion.drop(columns=['Total Time'], inplace=True)

# Extract relevant columns and handle call results and durations for Retention Agents
call_data_retention = call_results_retention[['Agent Name', 'Unique', 'Call Attempts']].copy()
duration_data_retention = duration_results_retention[['Agent Name', 'Total Time']].copy()

# Convert the "Total Time" to timedelta and handle formatting
duration_data_retention['Total Time'] = pd.to_timedelta(duration_data_retention['Total Time'], errors='coerce')
duration_data_retention['Duration'] = duration_data_retention['Total Time'].apply(lambda x: str(x).split()[-1] if pd.notnull(x) else '0')
duration_data_retention.drop(columns=['Total Time'], inplace=True)

# Merge the "Unique", "Call Attempts" and "Duration" columns with the agent data for Conversion Agents
merged_data_conversion = pd.merge(agents_data[agents_data['DEPARTMENT'] == 1].copy(), call_data_conversion, on='Agent Name', how='left')
merged_data_conversion = pd.merge(merged_data_conversion, duration_data_conversion, on='Agent Name', how='left')

# Merge the "Unique", "Call Attempts" and "Duration" columns with the agent data for Retention Agents
merged_data_retention = pd.merge(agents_data[agents_data['DEPARTMENT'] == 2].copy(), call_data_retention, on='Agent Name', how='left')
merged_data_retention = pd.merge(merged_data_retention, duration_data_retention, on='Agent Name', how='left')

# Fill missing values in "Unique", "Call Attempts", and "Duration" with 0
merged_data_conversion['Unique'] = merged_data_conversion['Unique'].fillna(0)
merged_data_conversion['Call Attempts'] = merged_data_conversion['Call Attempts'].fillna(0)
merged_data_conversion['Duration'] = merged_data_conversion['Duration'].fillna('0')
merged_data_conversion['Target'] = 0  # Default target

merged_data_retention['Unique'] = merged_data_retention['Unique'].fillna(0)
merged_data_retention['Call Attempts'] = merged_data_retention['Call Attempts'].fillna(0)
merged_data_retention['Duration'] = merged_data_retention['Duration'].fillna('0')
merged_data_retention['Target'] = 0  # Default target

# Define a function to calculate the Target as a percentage
def calculate_target(row):
    # Calculate percentage from unique calls
    unique_calls_percentage = (row['Unique'] // 3) * 1

    # Calculate percentage from duration
    duration_str = str(row['Duration'])
    if duration_str != '0':
        duration_parts = duration_str.split(':')
        if len(duration_parts) == 3:
            hours, minutes, seconds = map(int, duration_parts)
        elif len(duration_parts) == 2:
            hours = 0
            minutes, seconds = map(int, duration_parts)
        else:
            hours = 0
            minutes = 0
            seconds = int(duration_parts[0])
        total_minutes = hours * 60 + minutes + (seconds / 60)
        duration_percentage = (total_minutes // 1.5) * 1
    else:
        duration_percentage = 0

    # Combine the percentages
    return unique_calls_percentage + duration_percentage

# Apply the calculation function to each row
merged_data_conversion['Target'] = merged_data_conversion.apply(calculate_target, axis=1)
merged_data_retention['Target'] = merged_data_retention.apply(calculate_target, axis=1)

# Convert the Target to a numeric value for sorting
merged_data_conversion['Target Numeric'] = merged_data_conversion['Target']
merged_data_retention['Target Numeric'] = merged_data_retention['Target']

# Apply the custom sorting order
merged_data_conversion['Desk Order'] = merged_data_conversion['Desk'].map(desk_order_dict)
merged_data_retention['Desk Order'] = merged_data_retention['Desk'].map(desk_order_dict)

# Sort dataframes by Desk Order and then by Target
merged_data_conversion.sort_values(by=['Desk Order', 'Target Numeric'], ascending=[True, False], inplace=True)
merged_data_retention.sort_values(by=['Desk Order', 'Target Numeric'], ascending=[True, False], inplace=True)

# Drop unnecessary columns
merged_data_conversion.drop(columns=['Desk Order', 'Target Numeric', 'DEPARTMENT'], inplace=True)
merged_data_retention.drop(columns=['Desk Order', 'Target Numeric', 'DEPARTMENT'], inplace=True)

# Reorder columns for Conversion Agents
merged_data_conversion = merged_data_conversion[['Desk', 'Agent Name', 'Duration', 'Call Attempts', 'Unique', 'Target']]

# Reorder columns for Retention Agents
merged_data_retention = merged_data_retention[['Desk', 'Agent Name', 'Duration', 'Call Attempts', 'Unique', 'Target']]

# Convert 'Target' column to string with percentage format
merged_data_conversion['Target'] = merged_data_conversion['Target'].astype(int).astype(str) + '%'
merged_data_retention['Target'] = merged_data_retention['Target'].astype(int).astype(str) + '%'

# Create the new Excel file with updated data
file_path = 'Agent_Results.xlsx'
with pd.ExcelWriter(file_path) as writer:
    merged_data_conversion.to_excel(writer, sheet_name='Conversion Agents', index=False)
    merged_data_retention.to_excel(writer, sheet_name='Retention Agents', index=False)

# Load the Excel file with openpyxl
wb = load_workbook(file_path)

# Get the worksheets
ws_conversion = wb['Conversion Agents']
ws_retention = wb['Retention Agents']

# Define colors for each desk
desk_colors = {
    'Japanese': 'FFC7CE',  # Light red
    'Korea': 'FFEB9C',     # Light yellow
    'India Ajay': 'C6EFCE', # Light green
    'India Aarav': '9CBEFF', # Light blue
    'India': 'D9EAD3',     # Light greenish
    'English': 'EAD1DC',   # Light pink
    'French': 'DDEBF7',    # Light blue
    'German': 'F4CCCC',    # Light red
    'Spanish': 'FFF2CC',   # Light yellow
    'Portuguese': 'D9D2E9' # Light purple
}

# Define a thin border
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def color_and_border_rows(worksheet, df):
    for index, row in enumerate(df.itertuples(), start=2):  # Start at row 2 to skip header
        desk = getattr(row, 'Desk')
        fill_color = PatternFill(start_color=desk_colors.get(desk, 'FFFFFF'), end_color=desk_colors.get(desk, 'FFFFFF'), fill_type='solid')
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=index, column=col)
            cell.fill = fill_color
            cell.border = thin_border

        # Autofit column width (approximation)
        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            col_letter = col[0].column_letter
            worksheet.column_dimensions[col_letter].width = max_length + 2  # Add padding

# Apply colors and borders to both worksheets
color_and_border_rows(ws_conversion, merged_data_conversion)
color_and_border_rows(ws_retention, merged_data_retention)

# Save the changes
wb.save(file_path)

print("\nAgent_Results.xlsx has been generated with colored rows, borders, and sorted by Desk and Target.")











































